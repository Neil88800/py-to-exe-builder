import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import threading
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

class ExcelNameCounterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("統計工具 Pro")
        self.root.geometry("640x500")
        self.root.configure(bg="#0f0f1a")
        self.root.resizable(False, False)

        self.filename = None
        self.sample_filename = None
        self.save_path = tk.StringVar()
        self.output_filename = tk.StringVar(value="產出報表")

        self.set_theme()

        title = tk.Label(root, text="報表用統計工具", font=("Segoe UI", 20, "bold"),
                         fg="#00f0ff", bg="#0f0f1a")
        title.pack(pady=20)

        self.select_btn = tk.Button(root, text="📂 載入 Excel 檔案", command=self.select_file,
                                    **self.button_style(text_color="#000000"))
        self.select_btn.pack(pady=10)

        self.file_label = tk.Label(root, text="尚未選擇檔案", font=("Segoe UI", 12),
                                   fg="#aaaaaa", bg="#0f0f1a")
        self.file_label.pack()

        sample_btn = tk.Button(root, text="📄 載入樣板檔案", command=self.select_sample_file,
                               **self.button_style(small=True, text_color="#000000"))
        sample_btn.pack(pady=10)

        self.sample_label = tk.Label(root, text="尚未選擇樣板", font=("Segoe UI", 12),
                                     fg="#aaaaaa", bg="#0f0f1a")
        self.sample_label.pack()

        path_frame = tk.Frame(root, bg="#0f0f1a")
        path_frame.pack(pady=10)
        tk.Label(path_frame, text="📁 儲存路徑：", font=("Segoe UI", 12),
                 fg="#ffffff", bg="#0f0f1a").pack(side=tk.LEFT, padx=5)
        self.path_entry = tk.Entry(path_frame, textvariable=self.save_path, width=40, **self.entry_style())
        self.path_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(path_frame, text="選擇", command=self.select_save_path,
                  **self.button_style(small=True, text_color="#000000")).pack(side=tk.LEFT)

        filename_frame = tk.Frame(root, bg="#0f0f1a")
        filename_frame.pack(pady=5)
        tk.Label(filename_frame, text="🗒️ 檔名（不需輸入副檔名）：", font=("Segoe UI", 12),
                 fg="#ffffff", bg="#0f0f1a").pack(side=tk.LEFT, padx=5)
        self.name_entry = tk.Entry(filename_frame, textvariable=self.output_filename, **self.entry_style())
        self.name_entry.pack(side=tk.LEFT)

        self.run_btn = tk.Button(root, text="🚀 開始分析", command=self.run_process,
                                 **self.button_style(text_color="#000000"))
        self.run_btn.pack(pady=20)

        self.progress = ttk.Progressbar(root, length=440, mode='determinate', style="Custom.Horizontal.TProgressbar")
        self.progress.pack(pady=8)

        self.progress_label = tk.Label(root, text="", font=("Segoe UI", 11), fg="#00f0ff", bg="#0f0f1a")
        self.progress_label.pack()

        self.footer = tk.Label(root, text="🔧 Powered by Pandas & Tkinter | Abby 專用版", font=("Segoe UI", 10),
                               fg="#5555aa", bg="#0f0f1a")
        self.footer.pack(pady=12)

    def set_theme(self):
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Custom.Horizontal.TProgressbar",
                        troughcolor='#1a1a2e',
                        background='#00f0ff',
                        bordercolor='#0f0f1a',
                        lightcolor='#00e0dd',
                        darkcolor='#00e0dd',
                        thickness=12)

    def button_style(self, small=False, text_color="#ffffff"):
        return {
            "font": ("Segoe UI", 10 if small else 14, "bold"),
            "fg": text_color,
            "bg": "#0077b6",
            "activebackground": "#00b4d8",
            "activeforeground": text_color,
            "relief": "flat",
            "bd": 0,
            "padx": 14,
            "pady": 7,
        }

    def entry_style(self):
        return {
            "font": ("Segoe UI", 12),
            "fg": "#ffffff",
            "bg": "#2a2a40",
            "insertbackground": "#00f0ff",
            "relief": "flat",
            "highlightthickness": 1,
            "highlightbackground": "#444444",
            "highlightcolor": "#00f0ff",
        }

    def select_file(self):
        self.filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if self.filename:
            self.file_label.config(text=os.path.basename(self.filename))

    def select_sample_file(self):
        self.sample_filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if self.sample_filename:
            self.sample_label.config(text=os.path.basename(self.sample_filename))

    def select_save_path(self):
        directory = filedialog.askdirectory()
        if directory:
            self.save_path.set(directory)

    def run_process(self):
        if not self.filename or not self.sample_filename:
            messagebox.showwarning("⚠️ 錯誤", "請選擇 Excel 檔案與樣板檔案。")
            return
        if not self.save_path.get():
            messagebox.showwarning("⚠️ 錯誤", "請選擇儲存位置。")
            return
        if not self.output_filename.get().strip():
            messagebox.showwarning("⚠️ 錯誤", "請輸入輸出檔名。")
            return

        self.progress['value'] = 0
        self.progress_label.config(text="")
        threading.Thread(target=self.process_file).start()

    def process_file(self):
        try:
            start_time = time.time()
            self.update_progress(10, start_time)

            df = pd.read_excel(self.filename, sheet_name='工作表1')
            data = df.iloc[1:, 2].dropna().astype(str)
            names = data.str.extract(r'-(.+)$')[0].str.strip()
            summary = names.value_counts().reset_index()
            summary.columns = ['姓名', '次數']
            summary_total = summary['次數'].sum()
            self.update_progress(40, start_time)

            wb = load_workbook(self.sample_filename)
            ws = wb["工作表1"]

            for row in ws.iter_rows(min_row=3, max_row=999, min_col=2, max_col=3):
                for cell in row:
                    cell.value = None

            start_row = 3
            for i, (name, count) in enumerate(summary.values):
                ws.cell(row=start_row + i, column=2, value=name)
                ws.cell(row=start_row + i, column=3, value=count)
                cell_d = ws.cell(row=start_row + i, column=4, value="是■      否□")
                cell_d.font = Font(name="標楷體", size=16)
                cell_d.alignment = Alignment(horizontal='left', vertical='center')

            final_row = start_row + len(summary)

            thin_top = Border(top=Side(style='thin'))
            for col in range(1, 6):
                cell = ws.cell(row=final_row, column=col)
                cell.font = Font(name="標楷體", size=16)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_top

            ws.cell(row=final_row, column=1).value = "承辦人："
            ws.cell(row=final_row, column=4).value = "單位主管："

            if final_row + 1 <= ws.max_row:
                ws.delete_rows(final_row + 1, ws.max_row - final_row)

            for row in ws.iter_rows(min_row=2, max_row=final_row, min_col=1, max_col=4):
                for cell in row:
                    cell.font = Font(name="標楷體", size=14)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            if "工作表2" in wb.sheetnames:
                del wb["工作表2"]
            ws2 = wb.create_sheet("工作表2")
            ws2["B2"] = "姓名"
            ws2["C2"] = "次數"
            for i, (name, count) in enumerate(summary.values, start=3):
                ws2[f"B{i}"] = name
                ws2[f"C{i}"] = count
            ws2[f"B{i+1}"] = "總計"
            ws2[f"C{i+1}"] = summary_total

            for row in ws2.iter_rows(min_row=2, max_row=i+1, min_col=2, max_col=3):
                for cell in row:
                    cell.font = Font(name="標楷體", size=14)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            filename = self.output_filename.get().strip()
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            output_path = os.path.join(self.save_path.get(), filename)

            wb.save(output_path)
            self.update_progress(100, start_time)
            self.root.after(0, self.show_complete_message, output_path)
        except Exception as e:
            self.root.after(0, self.show_error_message, e)

    def update_progress(self, percent, start_time):
        self.progress['value'] = percent
        elapsed = time.time() - start_time
        if percent > 0:
            remaining = int((elapsed / percent) * (100 - percent))
            self.progress_label.config(text=f"進度：{percent}%　預估剩餘時間：{remaining} 秒")
        else:
            self.progress_label.config(text="")
        self.root.update_idletasks()

    def show_complete_message(self, output_path):
        messagebox.showinfo("✅ 完成", f"處理完成！結果已儲存於：\n{output_path}")

    def show_error_message(self, e):
        messagebox.showerror("❌ 錯誤", f"發生錯誤：\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelNameCounterApp(root)
    root.mainloop()
